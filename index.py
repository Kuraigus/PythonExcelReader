import pandas as pd
import math
import openpyxl

cidade = {}

file = "Mapa de Oportunidades - Brasíla - 11.09.2020.xlsb"
print("lendo planilha...")
sheet = pd.read_excel(file, engine='pyxlsb', sheet_name="BASE INFORMAÇÕES", skiprows=16, usecols=[
                      "BAIRRO / CIDADE", 'HPs CADASTRADOS', 'INSTALADOS', ' HP LIVRE'])

print("estruturando dados...")
for index, row in sheet.iterrows():
    if row["BAIRRO / CIDADE"] not in cidade:
        cidade[row["BAIRRO / CIDADE"]] = {}
    if "HPs CADASTRADOS" not in cidade[row["BAIRRO / CIDADE"]]:
        cidade[row["BAIRRO / CIDADE"]]["HPs CADASTRADOS"] = 0
    if "INSTALADOS" not in cidade[row["BAIRRO / CIDADE"]]:
        cidade[row["BAIRRO / CIDADE"]]["INSTALADOS"] = 0
    if "HP LIVRE" not in cidade[row["BAIRRO / CIDADE"]]:
        cidade[row["BAIRRO / CIDADE"]]["HP LIVRE"] = 0

    cidade[row["BAIRRO / CIDADE"]]["HPs CADASTRADOS"] += row["HPs CADASTRADOS"]
    if math.isnan(row["INSTALADOS"]) == False:
        cidade[row["BAIRRO / CIDADE"]]["INSTALADOS"] += row["INSTALADOS"]
    cidade[row["BAIRRO / CIDADE"]]["HP LIVRE"] += row[" HP LIVRE"]


for key, value in cidade.items():
    percentage = value["INSTALADOS"] * 100 / value["HPs CADASTRADOS"]
    cidade[key]["penetracao"] = "{:.2f}%".format(percentage)

cidade_sorted_cadastros = dict(
    sorted(cidade.items(), key=lambda k: k[1]['HPs CADASTRADOS'], reverse=True))

cidade_sorted_pen = dict(
    sorted(cidade.items(), key=lambda k: k[1]['penetracao']))


print("criando arquivo excel...")
# Create the workbook and sheet for Excel
workbook = openpyxl.Workbook()
new_sheet = workbook.active

row = 1
new_sheet.cell(row=row, column=1, value="Lista por cadastrados")

row += 1
new_sheet.cell(row=row, column=1, value="BAIRRO / CIDADE")
new_sheet.cell(row=row, column=2, value="HPs CADASTRADOS")
new_sheet.cell(row=row, column=3, value="INSTALADOS")
new_sheet.cell(row=row, column=4, value="HP LIVRE")
new_sheet.cell(row=row, column=5, value="PENETRAÇÃO")

row += 1
for key, values in cidade_sorted_cadastros.items():
    # Put the key in the first column for each key in the dictionary
    new_sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values.values():
        # Put the element in each adjacent column for each element in the tuple
        new_sheet.cell(row=row, column=column, value=element)
        column += 1
    row += 1

row += 2

new_sheet.cell(row=row, column=1, value="Lista por penetracao")

row += 1
new_sheet.cell(row=row, column=1, value="BAIRRO / CIDADE")
new_sheet.cell(row=row, column=2, value="HPs CADASTRADOS")
new_sheet.cell(row=row, column=3, value="INSTALADOS")
new_sheet.cell(row=row, column=4, value="HP LIVRE")
new_sheet.cell(row=row, column=5, value="PENETRAÇÃO")

row += 1
for key, values in cidade_sorted_pen.items():
    # Put the key in the first column for each key in the dictionary
    new_sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values.values():
        # Put the element in each adjacent column for each element in the tuple
        new_sheet.cell(row=row, column=column, value=element)
        column += 1
    row += 1


workbook.save(filename="feedback.xlsx")

print("Concluido !!")
